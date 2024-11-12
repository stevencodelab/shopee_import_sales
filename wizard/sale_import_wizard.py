from odoo import models, fields, api, _
from odoo.exceptions import UserError, ValidationError
import base64
import csv
import io
import logging
from datetime import datetime
import tempfile

_logger = logging.getLogger(__name__)

try:
    import pytz
except ImportError:
    pytz = None
    _logger.warning("pytz library is not installed. Timezone conversion may not be accurate.")

try:
    import xlrd
    import openpyxl
except ImportError:
    _logger.warning("xlrd or openpyxl library is not installed. Excel import may not work.")

class SaleImportWizard(models.TransientModel):
    _name = 'sale.import.wizard'
    _description = 'Sale Import Wizard'

    file_data = fields.Binary(string='File', required=True)
    filename = fields.Char(string='Filename')
    file_type = fields.Selection([
        ('csv', 'CSV File'),
        ('xls', 'XLS File'),
        ('xlsx', 'XLSX File')
    ], string='File Type', required=True, default='csv')
    marketplace_id = fields.Many2one('market.place', string='Marketplace', required=True)

    @api.onchange('filename')
    def _onchange_filename(self):
        if self.filename:
            file_extension = self.filename.split('.')[-1].lower()
            if file_extension in ['csv']:
                self.file_type = 'csv'
            elif file_extension in ['xls']:
                self.file_type = 'xls'
            elif file_extension in ['xlsx']:
                self.file_type = 'xlsx'

    def _parse_file(self):
        """Parse the uploaded file based on its type"""
        if not self.file_data:
            raise UserError(_("Please upload a file to import."))

        if self.file_type == 'csv':
            return self._parse_csv()
        elif self.file_type == 'xls':
            return self._parse_xls()
        elif self.file_type == 'xlsx':
            return self._parse_xlsx()
        else:
            raise UserError(_("Unsupported file type."))

    def _parse_csv(self):
        """Parse CSV file"""
        data = base64.b64decode(self.file_data)
        encodings = ['utf-8', 'iso-8859-1', 'windows-1252']
        for encoding in encodings:
            try:
                file_input = io.StringIO(data.decode(encoding))
                reader = csv.DictReader(file_input, delimiter=',')
                return list(reader)
            except UnicodeDecodeError:
                continue
        raise UserError(_("Unable to decode the CSV file. Please check the file encoding."))

    def _parse_xls(self):
        """Parse XLS file"""
        data = base64.b64decode(self.file_data)
        
        # Write the binary data to a temporary file
        with tempfile.NamedTemporaryFile(delete=False) as temp_file:
            temp_file.write(data)
            temp_file.seek(0)
            
            try:
                # Open the workbook
                book = xlrd.open_workbook(temp_file.name)
                sheet = book.sheet_by_index(0)
                
                # Get headers from the first row
                headers = [str(cell.value) for cell in sheet.row(0)]
                
                # Convert all rows to dictionaries
                result = []
                for row_idx in range(1, sheet.nrows):
                    row = {}
                    for col_idx, header in enumerate(headers):
                        cell_value = sheet.cell(row_idx, col_idx).value
                        # Convert numbers to strings to match CSV behavior
                        if isinstance(cell_value, float):
                            # Check if it's actually an integer
                            if cell_value.is_integer():
                                cell_value = str(int(cell_value))
                            else:
                                cell_value = str(cell_value)
                        row[header] = str(cell_value) if cell_value else ''
                    result.append(row)
                return result
            except Exception as e:
                raise UserError(_("Error reading XLS file: %s") % str(e))

    def _parse_xlsx(self):
        """Parse XLSX file"""
        data = base64.b64decode(self.file_data)
        
        # Write the binary data to a temporary file
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
            temp_file.write(data)
            temp_file.seek(0)
            
            try:
                # Load the workbook
                workbook = openpyxl.load_workbook(temp_file.name, data_only=True)
                sheet = workbook.active
                
                # Get headers from the first row
                headers = [str(cell.value) for cell in next(sheet.rows)]
                
                # Convert all rows to dictionaries
                result = []
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    row_dict = {}
                    for header, value in zip(headers, row):
                        # Convert numbers to strings to match CSV behavior
                        if isinstance(value, (int, float)):
                            if isinstance(value, float) and value.is_integer():
                                value = str(int(value))
                            else:
                                value = str(value)
                        row_dict[header] = str(value) if value is not None else ''
                    result.append(row_dict)
                return result
            except Exception as e:
                raise UserError(_("Error reading XLSX file: %s") % str(e))

    def _parse_datetime(self, date_string):
        """
        Parse date from string format to datetime
        """
        if not date_string:
            return False
        
        date_formats = [
            '%m/%d/%Y %H:%M',  # Format in your CSV: 9/21/2024 8:39
            '%m/%d/%Y',
            '%d/%m/%Y %H:%M',
            '%d/%m/%Y',
            '%Y-%m-%d %H:%M:%S',
            '%Y-%m-%d %H:%M',
            '%Y-%m-%d',
        ]

        for fmt in date_formats:
            try:
                return datetime.strptime(date_string.strip(), fmt)
            except ValueError:
                continue
    
        _logger.warning(f"Unable to parse date: {date_string}")
        return False

    def _parse_float(self, value):
        if not value:
            return 0.0
        try:
            # Menghapus pemisah ribuan dan mengganti pemisah desimal
            cleaned_value = value.replace('.', '').replace(',', '.')
            return float(cleaned_value)
        except ValueError:
            _logger.warning(f"Invalid float value: {value}")
            return 0.0

    def _get_or_create_partner(self, row):
        """
        Get or create partner based on username
        """
        Partner = self.env['res.partner']
        username = row.get('Username (Pembeli)')
        
        if not username:
            raise ValidationError(_("Username (Pembeli) is required to create or find a partner."))
        
        partner = Partner.search([('name', '=', username)], limit=1)
        if not partner:
            partner_vals = {
                'name': username,
                'phone': row.get('No. Telepon'),
                'street': row.get('Alamat Pengiriman'),
                'city': row.get('Kota/Kabupaten'),
                'state_id': self._get_state_id(row.get('Provinsi')),
            }
            partner = Partner.create(partner_vals)
        return partner
    
    def _get_state_id(self, state_name):
        """
        Get state ID based on name
        """
        State = self.env['res.country.state']
        state = State.search([('name', '=', state_name)], limit=1)
        return state.id if state else False

    def _get_or_create_product(self, row):
        """
        Get or create product based on SKU
        """
        Product = self.env['product.product']
        product = Product.search([('default_code', '=', row.get('Nomor Referensi SKU'))], limit=1)
        if not product:
            product = Product.create({
                'name': row.get('Nama Produk'),
                'default_code': row.get('Nomor Referensi SKU'),
                'list_price': self._parse_float(row.get('Harga Awal')),
                'weight': self._parse_float(row.get('Berat Produk')),
            })
        return product
    
    def _get_or_create_carrier(self, carrier_name):
        """
        Get or create delivery carrier based on name
        """
        Carrier = self.env['delivery.carrier']
        
        if not carrier_name:
            return False
        
        carrier = Carrier.search([('name', '=', carrier_name)], limit=1)
        if not carrier:
            carrier_vals = {
                'name': carrier_name,
                'delivery_type': 'fixed',  # Anda bisa menyesuaikan tipe delivery sesuai kebutuhan
                'product_id': self.env.ref('delivery.product_product_delivery').id,
            }
            carrier = Carrier.create(carrier_vals)
        return carrier    
    
    def _create_sale_order(self, row):
        """
        Create or update sale order based on CSV row data
        """
        SaleOrder = self.env['sale.order']
        order = SaleOrder.search([('nomor_pesanan', '=', row.get('No. Pesanan'))], limit=1)

        # Mendapatkan payment mode 'BC Online'
        payment_mode = self.env['account.payment.mode'].search([('name', '=', 'BC Online')], limit=1)
        if not payment_mode:
            raise ValidationError(_("Payment mode 'BC Online' not found in the system."))
        
        # # Mendapatkan automatic workflow 'Automatic'
        workflow = self.env['sale.workflow.process'].search([('name', '=', 'Automatic')], limit=1)
        if not workflow:
            raise ValidationError(_("Workflow 'Automatic' not found in the system."))

        partner = self._get_or_create_partner(row)
        carrier = self._get_or_create_carrier(row.get('Opsi Pengiriman'))
        
        order_vals = {
            'partner_id': partner.id,
            'nomor_pesanan': row.get('No. Pesanan'),
            'order_status': row.get('Status Pesanan'),
            'cancellation_return_status': row.get('Status Pembatalan/ Pengembalian'),
            'tracking_number': row.get('No. Resi'),
            'opsi_pengiriman': row.get('Opsi Pengiriman'),
            'carrier_id': carrier.id if carrier else False,
            'shipping_option': 'antar counter' if row.get('Antar ke counter/ pick-up') == 'Antar Ke Counter' else 'pickup',
            'must_ship_before': self._parse_datetime(row.get('Pesanan Harus Dikirimkan Sebelum (Menghindari keterlambatan)')),
            'order_creation_time': self._parse_datetime(row.get('Waktu Pesanan Dibuat')),
            'payment_time': self._parse_datetime(row.get('Waktu Pembayaran Dilakukan')),
            'payment_method': row.get('Metode Pembayaran'),
            'platform_discount': self._parse_float(row.get('Diskon Dari Shopee')),
            'cashback': self._parse_float(row.get('Cashback Koin')),
            'voucher_platform': self._parse_float(row.get('Voucher Ditanggung Shopee')),
            'package_discount': self._parse_float(row.get('Paket Diskon')),
            'package_discount_platform': self._parse_float(row.get('Paket Diskon (Diskon dari Shopee)')),
            'package_discount_seller': self._parse_float(row.get('Paket Diskon (Diskon dari Penjual)')),
            'coin_discount': self._parse_float(row.get('Potongan Koin Shopee')),
            'credit_card_discount': self._parse_float(row.get('Diskan Kartu Kredit')),
            'shipping_fee_paid_by_buyer': self._parse_float(row.get('Ongkos Kirim Dibayar oleh Pembeli')),
            'shipping_fee_discount': self._parse_float(row.get('Estimasi Potongan Biaya Pengiriman')),
            'return_shipping_fee': self._parse_float(row.get('Ongkos Kirim Pengembalian Barang')),
            'estimated_shipping_fee': self._parse_float(row.get('Perkiraan Ongkos Kirim')),
            'buyer_note': row.get('Catatan dari Pembeli'),
            'buyer_username': row.get('Username (Pembeli)'),
            'receiver_name': row.get('Nama Penerima'),
            'receiver_phone': row.get('No. Telepon'),
            'shipping_address': row.get('Alamat Pengiriman'),
            'city': row.get('Kota/Kabupaten'),
            'province': row.get('Provinsi'),
            'order_completion_time': self._parse_datetime(row.get('Waktu Pesanan Selesai')),
            'sale_marketplace': self.marketplace_id.id,
            # Menambahkan payment mode dan workflow
            'payment_mode_id': payment_mode.id,
            'workflow_process_id': workflow.id,
        }

        if order:
            order.write(order_vals)
        else:
            order = SaleOrder.create(order_vals)

        # Process order lines
        product = self._get_or_create_product(row)
        
        # Parse values from CSV
        original_price = self._parse_float(row.get('Harga Awal'))
        discounted_price = self._parse_float(row.get('Harga Setelah Diskon'))
        quantity = self._parse_float(row.get('Jumlah'))

        # Calculate discount percentage
        if original_price > 0:
            discount_percentage = ((original_price - discounted_price) / original_price) * 100
        else:
            discount_percentage = 0.0
        
        line_vals = {
            'order_id': order.id,
            'product_id': product.id,
            'parent_sku': row.get('SKU Induk'),
            'sku_reference': row.get('Nomor Referensi SKU'),
            'variation_name': row.get('Nama Variasi'),
            'original_price': original_price,
            'discounted_price': discounted_price,
            'returned_quantity': self._parse_float(row.get('Returned quantity', '0')),
            'product_uom_qty': quantity,
            'product_weight': self._parse_float(row.get('Berat Produk')),
            'total_weight': self._parse_float(row.get('Total Berat')),
            'discount': discount_percentage,
            'price_unit': original_price,
        }
        order.order_line = [(0, 0, line_vals)]
        
        return order
        
    def import_sales(self):
        """Import sales from the uploaded CSV file."""
        self.ensure_one()
        if not self.file_data:
            raise UserError(_("Please upload a file to import."))
        rows = self._parse_file()
        created_orders = self.env['sale.order']
        errors = []

        for index, row in enumerate(rows, start=1):
            try:
                order = self._create_sale_order(row)
                created_orders |= order
            except ValidationError as e:
                errors.append(f"Row {index}: Validation error - {str(e)}")
            except Exception as e:
                errors.append(f"Row {index}: Unexpected error - {str(e)}")
                _logger.exception("Error importing row %s: %s", index, str(e))
        
        if errors:
            raise UserError("\n".join(errors))
    
        return {
            'type': 'ir.actions.act_window',
            'name': _('Imported Sales Orders'),
            'res_model': 'sale.order',
            'view_mode': 'list,form',
            'domain': [('id', 'in', created_orders.ids)],
            'context': {'create': False},
        }